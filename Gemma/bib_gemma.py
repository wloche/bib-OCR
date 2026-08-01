#!/usr/bin/env python3
"""
bib_gemma.py — Detect runner/rider bib numbers in race photos using a local
vision-language model (Google's Gemma, served through Ollama).

Unlike plain OCR (see bib_ocr.py), a vision-language model can *reason* about
the image — telling a pinned bib number apart from a sponsor banner number,
handling partial occlusion, guessing at an angled/blurry digit from context,
etc. It should be noticeably more reliable than EasyOCR, at the cost of being
slower per photo and needing a heavier local runtime (Ollama + a multi-GB
model download) instead of a small pip package.

────────────────────────────────────────────────────────────────────────────
ONE-TIME SETUP
────────────────────────────────────────────────────────────────────────────
1. Install Ollama (the local model runtime):
     macOS:   brew install ollama        (or download from https://ollama.com)
     Then either open the Ollama app, or start the background server manually:
         ollama serve

2. Pull the vision model this script targets (~4GB download, one-time):
     ollama pull gemma4:e4b

   Other sizes work too (pass --model to use them):
     ollama pull gemma4:e2b     # smaller/faster, less accurate, ~3GB
     ollama pull gemma4:12b     # bigger/slower, more accurate

3. Install the Python client:
     pip install ollama

────────────────────────────────────────────────────────────────────────────
USAGE
────────────────────────────────────────────────────────────────────────────
    python bib_gemma.py --input /path/to/photos --output bib_numbers_gemma.csv
    python bib_gemma.py --input /path/to/photos --output bib_numbers_gemma.xlsx --recursive
    python bib_gemma.py --input /path/to/photos --model gemma4:12b --output out.csv
    python bib_gemma.py --input /path/to/photos --limit 10 --output sample.csv
    python bib_gemma.py --input /path/to/photos --quiet --output out.csv
    python bib_gemma.py --input /path/to/photos --bib-range 1000-1699 --output out.csv
    python bib_gemma.py --input /path/to/cycling --race-type cyclists --output out.csv
    python bib_gemma.py --input /path/to/photos --teams "xXx, 606" --output out.csv

OUTPUT
    A CSV (or XLSX, if the output filename ends in .xlsx) with one row per photo:
        filename, bib_numbers, colors, [team,] notes
    Multiple bibs in one photo are separated with semicolons: "482;217"
    "none" means there was no bib in the photo at all.
    "illegible" means a bib was visible but not readable — on its own, or
    alongside numbers ("852;illegible") when only some bibs could be read.
    "colors" is what the model reports seeing for each bib, in the same order
    ("black on white; unclear"). It is recorded, never used to filter — see
    FILTERING below for why.
    "team" only appears when --teams is used; see TEAMS below.

TEAMS
    --teams adds a "team" column holding the club/team name on each person's
    kit, in the same order as their bib ("xXx; 606"):

        --teams                 read whatever team name is on the kit
        --teams "xXx, 606"      as above, but match against the teams you know
                                are competing

    With a list, a case-insensitive match is normalised to your spelling ("xxx"
    -> "xXx") so the column groups cleanly. A team that isn't on the list is
    kept verbatim rather than rewritten or dropped: a name the model saw that
    you didn't list is information, not an error to hide. Expect "unknown" a lot
    — the prompt asks for it explicitly rather than letting the model infer a
    team from kit colours, and on small/blurry kit that is the honest answer.
    Note the list is given to the model as context, so the same priming caution
    as for colors applies: treat the column as a lead to verify, not a fact.

RACE TYPE
    --race-type runners|cyclists|mixed tunes what the model is told to look for.
    This matters more than it sounds: the default "mixed" prompt has to cover
    both disciplines, so it permits numbers "printed directly on a jersey" —
    harmless for runners, but in cycling frames full of sponsor lettering it is
    the main source of fabricated readings. On the 566-photo reference set every
    single invent-a-number-from-nothing error fell in the cycling half.

        --race-type runners    bib is a card pinned to the front torso; almost
                               every runner has one; ignore clocks and signage
        --race-type cyclists   small number on hip/lower back or a frame plate;
                               "none" is often correct; sponsor text is not a bib

    Run each block of photos separately with the matching race type when you
    can. Note that cyclists at these events do wear numbers — they are just
    small and often unreadable — so the cyclists prompt is written to reach for
    "illegible", not to assume there is nothing there.

FILTERING
    The known shape of a race's numbering is a far stronger signal than
    anything the model can self-report. On the 566-photo reference set, 134 of
    151 real bibs were 4-digit in the 1000-1699 band, while most fabricated
    readings were 1-3 digits. So:

        --min-digits 3        drop readings shorter than 3 digits
        --bib-range 1000-1699 drop readings outside the race's number range

    Both are applied after parsing, per reading, and discarded numbers are
    recorded in the notes column ("filtered out: 45, 7") so that a too-strict
    filter is visible in the output rather than silently swallowing real bibs.
    An "illegible" marker is never filtered — there is no number to judge.

    Colors are deliberately *not* used as a filter. Race lighting (backlight,
    shade, motion blur, mixed white balance) shifts apparent color enough that
    filtering on it would cost real detections, and a small vision model tends
    to agree with whatever color scheme you suggest to it. Recording the color
    instead lets you check offline whether it actually correlates with
    correctness, without paying for another multi-hour run.

PROGRESS
    Each photo prints its own line as it finishes, with the time that photo
    took, the total elapsed time, and an ETA for the whole run based on the
    average per-photo time so far:
        [12/566] _5D_0012.JPG ... 129;1171 [8.4s, elapsed 1m 41s, ETA 1h 17m 30s]
    The final summary reports the total wall-clock time and the average time
    per photo, which is handy for estimating a bigger batch or comparing models.

    Use --limit N for a quick trial run over just the first N photos — the
    script says up front how many of the photos it found are being skipped,
    and repeats that reminder at the end so a partial output file isn't
    mistaken for a complete one.

    --quiet silences all of that (progress lines, ETA, the --limit notice and
    the final summary) and leaves only errors and warnings, on stderr: a
    per-photo "parse_error: <file>: <raw response>" line for each photo that
    failed, plus the usual fatal errors. Handy for cron jobs and scripts —
    redirect stderr to a log and no news is good news. The output CSV/XLSX is
    identical either way.

NOTES
    - This calls a local Ollama server (default http://localhost:11434) — no
      internet connection or API key is used once the model is downloaded.
    - Runs entirely on your machine's CPU/GPU. Without a GPU, expect roughly
      a few seconds to tens of seconds per photo depending on model size —
      566 photos could take anywhere from ~30 minutes to a few hours on the
      e4b model on CPU alone. A GPU will be dramatically faster.
    - If a response can't be parsed (model didn't follow the expected format),
      the row is marked "parse_error" with the raw response in notes so you
      can inspect what happened, rather than silently guessing.
"""

import argparse
import base64
import csv
import re
import sys
import time
from pathlib import Path

# Sentinel for "a bib is there, but I can't read it" — distinct from "none".
# Without this the model has to choose between a number and denying the bib
# exists, and it reliably picks a fabricated number.
ILLEGIBLE = "illegible"

# What to look for, per discipline. Only this opening paragraph changes; the
# response format below is shared so parsing is identical across race types.
#
# The "mixed" text has to accommodate both disciplines at once, which is why it
# ends up permitting jersey-printed numbers — harmless for runners, but the main
# source of fabricated readings in cycling frames full of sponsor lettering.
# Naming the discipline lets each variant be strict instead of accommodating.
PROMPT_LEADS = {
    "mixed": (
        "You are looking at a photo from a running/cycling race. Look carefully at every "
        "person in the frame for a race bib number — a numbered card, tag, or plate pinned "
        "to someone's chest, waist, or back (or printed directly on a jersey)."
    ),
    "runners": (
        "You are looking at a photo from a running race. Look carefully at every runner in "
        "the frame for a race bib: a numbered paper or plastic card pinned to the front of "
        "the torso — chest or stomach — with large printed digits. A few runners pin theirs "
        "to the back instead. Nearly every runner in an official race wears one, so if you "
        "can see a runner's front torso clearly, there is usually a bib to find.\n\n"
        "Do NOT report numbers from timing clocks, road signs, sponsor banners, vehicles, "
        "or clothing brand graphics — only a pinned race bib counts."
    ),
    "cyclists": (
        "You are looking at a photo from a cycling race. Riders' numbers are small and easy "
        "to miss: usually a square number pinned to the lower back, hip, or side of the "
        "jersey, sometimes a plate mounted on the bike's frame or handlebars. They are often "
        "creased, curved around the body, or hidden by the rider's arm or riding position.\n\n"
        "Two things to be careful about:\n"
        "- Many riders in these photos wear no visible race number at all. \"none\" is a "
        "common and correct answer here.\n"
        "- Team kit is covered in sponsor lettering, brand names and numbers that are NOT "
        "race bibs. Do not report sponsor text, jersey logos, bike model numbers, or numbers "
        "on barriers, banners and vehicles. Only a pinned race number or a mounted number "
        "plate counts."
    ),
}

RESPONSE_FORMAT = """
Respond in EXACTLY this format, nothing else:

BIBS: <comma-separated numbers; write "illegible" in place of a number for a bib you can see but cannot actually read; write "none" if there is no bib at all>
COLORS: <for each entry in BIBS, "<digit color> on <background color>", comma-separated in the same order — leave blank if BIBS is none>
{team_line}NOTES: <one short phrase — leave blank if nothing notable, e.g. "back bib" or "partially obscured">

Do not guess at digits you cannot make out. "illegible" is a better answer than a number you are unsure of, and "none" is a better answer than a bib that isn't there.

Report the colors you actually see. Do not guess at a colour scheme you expect a race to use — if a bib's colors aren't clear, write "unclear" for that bib.
{team_guidance}
Examples of valid responses:
BIBS: 482
COLORS: black on white
{team_example_1}NOTES:

BIBS: 482, 217, 90
COLORS: black on white, black on white, unclear
{team_example_2}NOTES: one bib partially cropped at frame edge

BIBS: 852, illegible
COLORS: black on white, unclear
{team_example_3}NOTES: second bib too small to read

BIBS: none
COLORS:
{team_example_4}NOTES: no pinned bib visible
"""

# Asked for only when --teams is used, so the default prompt stays as short as
# it was — every extra instruction is a chance for a small model to drift.
TEAM_LINE = ('TEAM: <for each entry in BIBS, the team or club name on that person\'s kit, '
             'or "unknown" — comma-separated in the same order>\n')

TEAM_GUIDANCE_FREE = """
For TEAM, read the team or club name printed on the person's jersey or kit. Write "unknown" if there is no team name visible or you cannot read it — do not infer a team from kit colours alone.
"""

TEAM_GUIDANCE_KNOWN = """
For TEAM, read the team or club name printed on the person's jersey or kit. These teams are known to be competing:
{teams}
If the kit matches one of them, use that exact name. If it clearly shows a different team, write the name you see. Write "unknown" if there is no team name visible or you cannot read it — do not pick a name from the list just because it is on the list, and do not infer a team from kit colours alone.
"""


def build_prompt(race_type: str, teams=None, ask_team: bool = False) -> str:
    """Assemble the prompt. teams is an optional list of known team names."""
    if not ask_team:
        fields = {"team_line": "", "team_guidance": "",
                  "team_example_1": "", "team_example_2": "",
                  "team_example_3": "", "team_example_4": ""}
    else:
        if teams:
            guidance = TEAM_GUIDANCE_KNOWN.format(teams="\n".join(f"- {t}" for t in teams))
            named = teams[0]
            second = teams[1] if len(teams) > 1 else "unknown"
        else:
            guidance = TEAM_GUIDANCE_FREE
            named, second = "Riverside Runners", "unknown"
        fields = {
            "team_line": TEAM_LINE,
            "team_guidance": guidance,
            "team_example_1": f"TEAM: {named}\n",
            "team_example_2": f"TEAM: {named}, {second}, unknown\n",
            "team_example_3": f"TEAM: {named}, unknown\n",
            "team_example_4": "TEAM:\n",
        }
    return PROMPT_LEADS[race_type] + "\n" + RESPONSE_FORMAT.format(**fields)


# [ \t]* rather than \s* so an empty field can't swallow the following line —
# "COLORS:\nNOTES: foo" must yield an empty colors field, not "NOTES: foo".
BIBS_RE = re.compile(r"BIBS:[ \t]*(.+)", re.IGNORECASE)
COLORS_RE = re.compile(r"COLORS?:[ \t]*(.*)", re.IGNORECASE)
TEAM_RE = re.compile(r"TEAMS?:[ \t]*(.*)", re.IGNORECASE)
NOTES_RE = re.compile(r"NOTES:[ \t]*(.*)", re.IGNORECASE)
ILLEGIBLE_RE = re.compile(r"illegible|unreadable|can'?t read|not legible", re.IGNORECASE)


def format_duration(seconds: float) -> str:
    """Human-readable duration: '42s', '3m 07s', '1h 24m 09s'."""
    seconds = int(round(max(seconds, 0)))
    hours, rem = divmod(seconds, 3600)
    minutes, secs = divmod(rem, 60)
    if hours:
        return f"{hours}h {minutes:02d}m {secs:02d}s"
    if minutes:
        return f"{minutes}m {secs:02d}s"
    return f"{secs}s"


def find_images(input_dir: Path, recursive: bool, extensions):
    pattern = "**/*" if recursive else "*"
    files = []
    for ext in extensions:
        files.extend(input_dir.glob(f"{pattern}.{ext}"))
        files.extend(input_dir.glob(f"{pattern}.{ext.upper()}"))
    return sorted(set(files))


def _field(text: str, pattern) -> str:
    match = pattern.search(text)
    return match.group(1).strip() if match else ""


def _aligned(parts, idx):
    """The per-bib value at position idx.

    The model is asked for one entry per bib in the same order as BIBS. If it
    gave a single value for several bibs, reuse it; if it gave none, blank.
    """
    if idx < len(parts):
        return parts[idx]
    return parts[0] if len(parts) == 1 else ""


def canonical_team(value: str, known_teams):
    """Match a reported team against the known list, case-insensitively.

    Returns the list's own spelling on a match ("xxx" -> "xXx") so the column
    groups cleanly. Anything unmatched is kept verbatim rather than rewritten or
    dropped — a team the model saw that isn't on the list is information, not an
    error to hide.
    """
    if not value or not known_teams:
        return value
    folded = value.strip().lower()
    for team in known_teams:
        if folded == team.lower():
            return team
    return value


def parse_response(text: str, known_teams=None):
    """Pull bib numbers, colors, teams and notes out of the model's reply.

    Returns (detections, notes, ok). Each detection is a dict:
        {"bib": "482" or "illegible", "color": "black on white", "team": "xXx"}
    Keeping the three together avoids the parallel-list alignment bugs that
    per-bib fields otherwise invite. An empty list means the model reported no
    bib at all. On a parse failure, detections is None and notes holds the raw
    reply.
    """
    bibs_match = BIBS_RE.search(text)
    if not bibs_match:
        return None, text.strip()[:200], False

    raw_bibs = bibs_match.group(1).strip()
    notes = _field(text, NOTES_RE)

    def split_field(pattern):
        raw = _field(text, pattern)
        return [p.strip() for p in raw.split(",")] if raw else []

    color_parts = split_field(COLORS_RE)
    team_parts = split_field(TEAM_RE)

    if raw_bibs.lower().startswith("none"):
        return [], notes, True

    # Walk the comma-separated entries so an "illegible" marker survives next to
    # real numbers ("852, illegible"). Per-bib fields align on the entry index,
    # which is the ordering the model was asked to use.
    entries = []
    for idx, part in enumerate(raw_bibs.split(",")):
        part = part.strip()
        if not part:
            continue
        nums = re.findall(r"\d+", part)
        if nums:
            entries.extend((n, idx) for n in nums)
        elif ILLEGIBLE_RE.search(part):
            entries.append((ILLEGIBLE, idx))

    # Deduplicate, keeping the fields given for each entry's first occurrence.
    # Repeated "illegible" markers collapse to one: the number of unreadable
    # bibs in a frame isn't something a small model reports reliably.
    seen = set()
    detections = []
    for value, idx in entries:
        if value in seen:
            continue
        seen.add(value)
        team = _aligned(team_parts, idx)
        detections.append({
            "bib": value,
            "color": _aligned(color_parts, idx),
            "team": canonical_team(team, known_teams),
        })

    return detections, notes, True


def parse_bib_range(spec: str):
    """Parse a '1000-1699' range spec into (low, high). Raises ValueError."""
    parts = spec.replace("–", "-").split("-")
    if len(parts) != 2 or not all(p.strip().isdigit() for p in parts):
        raise ValueError(f"expected LOW-HIGH with whole numbers, got '{spec}'")
    low, high = int(parts[0]), int(parts[1])
    if low > high:
        raise ValueError(f"low bound {low} is greater than high bound {high}")
    return low, high


def apply_filters(detections, min_digits, bib_range):
    """Drop implausible readings. Returns (kept_detections, dropped_bibs).

    The known shape of a race's bib numbers is a much stronger signal than
    anything the model can tell us: on the reference set, 134 of 151 real bibs
    were 4-digit, while most fabricated readings were 1-3 digits. Dropped
    numbers are reported rather than silently discarded.
    """
    kept, dropped = [], []
    for d in detections:
        n = d["bib"]
        if n == ILLEGIBLE:
            # Not a reading, so there's nothing to judge plausible — keep it.
            kept.append(d)
        elif min_digits and len(n) < min_digits:
            dropped.append(n)
        elif bib_range and not (bib_range[0] <= int(n) <= bib_range[1]):
            dropped.append(n)
        else:
            kept.append(d)
    return kept, dropped


# Column metadata, keyed by the CSV header name. The "team" column is only
# present when --teams is used, so both writers take the column list.
XLSX_HEADERS = {
    "filename": "Photo Filename",
    "bib_numbers": "Bib Number(s)",
    "colors": "Colors",
    "team": "Team",
    "notes": "Notes",
}
XLSX_WIDTHS = {"filename": 22, "bib_numbers": 30, "colors": 28, "team": 22, "notes": 55}


def write_csv(rows, output_path: Path, columns):
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(columns)
        writer.writerows(rows)


def write_xlsx(rows, output_path: Path, columns):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is required for .xlsx output. Install it with:\n    pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bib Numbers"

    ncols = len(columns)
    bib_col = columns.index("bib_numbers") + 1

    ws.append([XLSX_HEADERS.get(c, c.title()) for c in columns])
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for col in range(1, ncols + 1):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial", size=10)
    none_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    illegible_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        for c in range(1, ncols + 1):
            ws.cell(row=r, column=c).font = body_font
        cell = ws.cell(row=r, column=bib_col)
        val = str(cell.value or "").strip().lower()
        if val == "none":
            cell.fill = none_fill
        elif val in ("parse_error", "error"):
            cell.fill = error_fill
        elif ILLEGIBLE in val:
            # Amber: worth a manual look, unlike a confident "none".
            cell.fill = illegible_fill

    for i, name in enumerate(columns, start=1):
        ws.column_dimensions[get_column_letter(i)].width = XLSX_WIDTHS.get(name, 20)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{ws.max_row}"
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Detect bib numbers in race photos using a local Gemma vision model via Ollama."
    )
    parser.add_argument("--input", required=True, help="Folder containing JPG photos")
    parser.add_argument("--output", default="bib_numbers_gemma.csv",
                         help="Output path (.csv or .xlsx). Default: bib_numbers_gemma.csv")
    parser.add_argument("--recursive", action="store_true", help="Also search subfolders")
    parser.add_argument("--model", default="gemma4:e4b",
                         help="Ollama model tag to use. Default: gemma4:e4b")
    parser.add_argument("--host", default="http://localhost:11434",
                         help="Ollama server URL. Default: http://localhost:11434")
    parser.add_argument("--retries", type=int, default=1,
                         help="Retries if the model's response can't be parsed. Default: 1")
    parser.add_argument("--limit", type=int, default=None,
                         help="Only process the first N photos found (useful for a quick "
                              "test run or benchmarking a model). Default: all of them")
    parser.add_argument("--quiet", action="store_true",
                         help="Suppress progress, ETA and summary output — report only "
                              "errors and warnings (on stderr). Useful for cron/scripts.")
    parser.add_argument("--min-digits", type=int, default=None, metavar="N",
                         help="Discard readings with fewer than N digits. Most fabricated "
                              "readings are 1-3 digits, so --min-digits 3 (or 4) cuts them.")
    parser.add_argument("--bib-range", default=None, metavar="LOW-HIGH",
                         help="Discard readings outside this numeric range, e.g. 1000-1699. "
                              "The strongest filter available if you know the race's numbering.")
    parser.add_argument("--teams", nargs="?", const="", default=None, metavar="LIST",
                         help="Also report each person's team in a 'team' column. Pass a "
                              "comma-separated list of the teams competing (e.g. "
                              "--teams 'xXx, 606') to have those names matched and spelled "
                              "consistently, or pass --teams alone to read whatever is on "
                              "the kit. Off by default.")
    parser.add_argument("--race-type", choices=sorted(PROMPT_LEADS), default="mixed",
                         help="Tailor the prompt to the discipline: 'runners' (bib pinned to "
                              "the front torso) or 'cyclists' (small number on hip/lower back, "
                              "sponsor lettering is not a bib). Default: mixed")
    args = parser.parse_args()

    def info(*parts, **kwargs):
        """Print normal progress output, unless --quiet was passed."""
        if not args.quiet:
            print(*parts, **kwargs)

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        print(f"Error: '{input_dir}' is not a directory", file=sys.stderr)
        sys.exit(1)

    if args.limit is not None and args.limit < 1:
        print("Error: --limit must be 1 or greater", file=sys.stderr)
        sys.exit(1)

    if args.min_digits is not None and args.min_digits < 1:
        print("Error: --min-digits must be 1 or greater", file=sys.stderr)
        sys.exit(1)

    ask_team = args.teams is not None
    known_teams = [t.strip() for t in args.teams.split(",") if t.strip()] if args.teams else []
    if ask_team and args.teams and not known_teams:
        print(f"Error: --teams was given '{args.teams}' but no team names could be read "
              f"from it. Use a comma-separated list, e.g. --teams 'xXx, 606'", file=sys.stderr)
        sys.exit(1)

    bib_range = None
    if args.bib_range:
        try:
            bib_range = parse_bib_range(args.bib_range)
        except ValueError as e:
            print(f"Error: bad --bib-range ({e})", file=sys.stderr)
            sys.exit(1)

    try:
        import ollama
    except ImportError:
        print("The 'ollama' Python package is not installed. Install it with:\n    pip install ollama",
              file=sys.stderr)
        sys.exit(1)

    client = ollama.Client(host=args.host)

    # Fail fast with a clear message if Ollama isn't reachable or the model isn't pulled.
    try:
        available = {m["model"] for m in client.list().get("models", [])}
    except Exception as e:
        print(f"Could not reach Ollama at {args.host} ({e}).\n"
              f"Make sure Ollama is running (try: ollama serve) and try again.", file=sys.stderr)
        sys.exit(1)

    if available and not any(args.model in m or m.startswith(args.model) for m in available):
        print(f"Warning: '{args.model}' doesn't appear in `ollama list`. "
              f"If this fails, run: ollama pull {args.model}", file=sys.stderr)

    images = find_images(input_dir, args.recursive, extensions=["jpg", "jpeg"])
    if not images:
        print(f"No JPG files found in '{input_dir}'"
              f"{' (searched recursively)' if args.recursive else ''}.", file=sys.stderr)
        sys.exit(1)

    total_found = len(images)
    if args.limit is not None and args.limit < total_found:
        images = images[:args.limit]
        info(f"NOTE: --limit {args.limit} is set — only the first {args.limit} of "
             f"{total_found} photo(s) found will be processed. The output file will "
             f"cover those {args.limit} photo(s) only.")

    prompt = build_prompt(args.race_type, known_teams, ask_team)
    columns = ["filename", "bib_numbers", "colors"] + (["team"] if ask_team else []) + ["notes"]

    info(f"Processing {len(images)} photo(s). Using model '{args.model}' via Ollama at {args.host}.")
    info(f"Prompt tuned for: {args.race_type}")
    if ask_team:
        info(f"Also reporting team per bib"
             + (f", matched against: {', '.join(known_teams)}" if known_teams
                else " (reading whatever is on the kit)"))
    if args.min_digits or bib_range:
        criteria = []
        if args.min_digits:
            criteria.append(f"at least {args.min_digits} digit(s)")
        if bib_range:
            criteria.append(f"within {bib_range[0]}-{bib_range[1]}")
        info(f"Filtering readings: keeping only numbers {' and '.join(criteria)}. "
             f"Discarded numbers are listed in the notes column.")

    rows = []
    filtered_count = 0
    team_photos = 0
    run_start = time.monotonic()
    for i, img_path in enumerate(images, 1):
        # No newline yet: the timing/ETA is appended once this photo is done.
        info(f"[{i}/{len(images)}] {img_path.name} ... ", end="", flush=True)
        photo_start = time.monotonic()

        detections, notes, ok = None, "", False
        last_raw = ""
        for attempt in range(args.retries + 1):
            try:
                response = client.chat(
                    model=args.model,
                    messages=[{
                        "role": "user",
                        "content": prompt,
                        "images": [str(img_path)],
                    }],
                )
                text = response["message"]["content"]
                last_raw = text
                detections, notes, ok = parse_response(text, known_teams)
                if ok:
                    break
            except Exception as e:
                last_raw = f"request failed: {e}"
                ok = False

        if ok:
            dropped = []
            if args.min_digits or bib_range:
                detections, dropped = apply_filters(detections, args.min_digits, bib_range)
                filtered_count += len(dropped)
            if dropped:
                # Keep the discarded readings visible so a too-strict filter is
                # obvious from the output file, not silently invisible.
                note_parts = [p for p in (notes, f"filtered out: {', '.join(dropped)}") if p]
                notes = " | ".join(note_parts)
            summary = ";".join(d["bib"] for d in detections) if detections else "none"
            row = [img_path.name, summary,
                   "; ".join(d["color"] for d in detections if d["color"])]
            if ask_team:
                row.append("; ".join(d["team"] for d in detections if d["team"]))
                team_photos += 1 if any(
                    d["team"] and d["team"].lower() != "unknown" for d in detections) else 0
            row.append(notes)
            rows.append(row)
        else:
            detail = last_raw.replace("\n", " ")[:200]
            summary = "parse_error"
            rows.append([img_path.name, summary, ""] + ([""] if ask_team else []) + [detail])
            if args.quiet:
                # The inline progress line is hidden, so failures still need a voice.
                print(f"parse_error: {img_path.name}: {detail}", file=sys.stderr)

        # ETA from the average time per photo so far, which smooths out the
        # variation between a quick photo and a slow, crowded one.
        photo_secs = time.monotonic() - photo_start
        elapsed = time.monotonic() - run_start
        remaining = len(images) - i
        if remaining:
            eta = (elapsed / i) * remaining
            info(f"{summary} [{photo_secs:.1f}s, elapsed {format_duration(elapsed)}, "
                 f"ETA {format_duration(eta)}]")
        else:
            info(f"{summary} [{photo_secs:.1f}s, elapsed {format_duration(elapsed)}]")

    total_secs = time.monotonic() - run_start

    output_path = Path(args.output)
    if output_path.suffix.lower() == ".xlsx":
        write_xlsx(rows, output_path, columns)
    else:
        write_csv(rows, output_path, columns)

    # A photo counts as "read" only if a number came out of it; a bib the model
    # saw but couldn't read is reported separately rather than as a detection.
    detected = sum(1 for r in rows if re.search(r"\d", r[1]))
    illegible_only = sum(1 for r in rows if r[1] == ILLEGIBLE)
    errors = sum(1 for r in rows if r[1] == "parse_error")
    skipped = total_found - len(rows)
    info(f"\nDone. {detected}/{len(rows)} photos yielded a bib number "
         f"({illegible_only} bib seen but unreadable, {errors} parse errors). "
         f"Wrote results to {output_path}")
    if ask_team:
        info(f"A team was identified on {team_photos}/{len(rows)} photos "
             f"(the rest are blank or 'unknown').")
    if filtered_count:
        info(f"Filters discarded {filtered_count} reading(s) as implausible "
             f"(see the notes column for which).")
    if skipped:
        info(f"Reminder: --limit was set, so {skipped} of the {total_found} photo(s) "
             f"found were not processed.")
    info(f"Total time: {format_duration(total_secs)} "
         f"({total_secs / len(rows):.1f}s per photo average)")


if __name__ == "__main__":
    main()
