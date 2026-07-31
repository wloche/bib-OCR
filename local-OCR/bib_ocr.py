#!/usr/bin/env python3
"""
bib_ocr.py — Detect runner/rider bib numbers in race photos using local, offline OCR.

Uses EasyOCR (a free, local deep-learning OCR engine) to find numeric text in each
photo and guesses which readings are bib numbers based on digit count.

INSTALL (one-time):
    pip install easyocr

    EasyOCR pulls in PyTorch automatically. The first run also downloads its
    recognition model (~100MB) and caches it locally — after that everything
    runs fully offline, no API key and no per-image cost.

USAGE:
    python bib_ocr.py --input /path/to/photos --output bib_numbers.csv
    python bib_ocr.py --input /path/to/photos --output bib_numbers.csv --recursive
    python bib_ocr.py --input /path/to/photos --output bib_numbers.xlsx --gpu

OUTPUT:
    A CSV (or XLSX, if the output filename ends in .xlsx) with one row per photo:
        filename, bib_numbers, notes
    - Multiple bibs in one photo are separated with semicolons: "482;217"
    - "none" means no plausible bib number was detected above the confidence threshold.

TUNING:
    Bib numbers are typically 2-5 digits. If you're getting junk numbers (from
    sponsor banners, street signs, etc.) or missing real bibs, adjust:
        --min-conf     raise to be stricter about OCR confidence (default 0.35)
        --min-digits / --max-digits   narrow the expected bib length

LIMITATIONS:
    Local OCR is much less accurate than a vision-capable AI model on angled,
    small, blurry, or partially obscured bibs (common in action race photography).
    Expect to need manual review/cleanup of the results, especially for "none"
    rows that might actually contain an unreadable bib.
"""

import argparse
import csv
import re
import sys
from pathlib import Path


def find_images(input_dir: Path, recursive: bool, extensions):
    pattern = "**/*" if recursive else "*"
    files = []
    for ext in extensions:
        files.extend(input_dir.glob(f"{pattern}.{ext}"))
        files.extend(input_dir.glob(f"{pattern}.{ext.upper()}"))
    return sorted(set(files))


def extract_bib(text: str, min_digits: int, max_digits: int):
    """Pull a digit run out of an OCR text token if it looks bib-number-shaped."""
    cleaned = re.sub(r"[^0-9]", "", text)
    if not cleaned:
        return None
    if min_digits <= len(cleaned) <= max_digits:
        return cleaned
    return None


def write_csv(rows, output_path: Path):
    with output_path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["filename", "bib_numbers", "notes"])
        writer.writerows(rows)


def write_xlsx(rows, output_path: Path):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("openpyxl is required for .xlsx output. Install it with:\n    pip install openpyxl",
              file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bib Numbers"

    header = ["Photo Filename", "Bib Number(s)", "Notes"]
    ws.append(header)
    for row in rows:
        ws.append(row)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")

    body_font = Font(name="Arial", size=10)
    none_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    for r in range(2, ws.max_row + 1):
        for c in range(1, 4):
            ws.cell(row=r, column=c).font = body_font
        if str(ws.cell(row=r, column=2).value or "").strip().lower() == "none":
            ws.cell(row=r, column=2).fill = none_fill

    widths = [22, 30, 55]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:C{ws.max_row}"

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(
        description="Detect bib numbers in race photos using local, offline OCR (EasyOCR)."
    )
    parser.add_argument("--input", required=True, help="Folder containing JPG photos")
    parser.add_argument("--output", default="bib_numbers.csv",
                         help="Output path (.csv or .xlsx). Default: bib_numbers.csv")
    parser.add_argument("--recursive", action="store_true", help="Also search subfolders")
    parser.add_argument("--min-conf", type=float, default=0.35,
                         help="Minimum OCR confidence (0-1) to keep a reading. Default: 0.35")
    parser.add_argument("--min-digits", type=int, default=2,
                         help="Minimum digit count to treat text as a bib number. Default: 2")
    parser.add_argument("--max-digits", type=int, default=5,
                         help="Maximum digit count to treat text as a bib number. Default: 5")
    parser.add_argument("--gpu", action="store_true",
                         help="Use a CUDA GPU if available (much faster; requires GPU torch build)")
    parser.add_argument("--languages", default="en",
                         help="Comma-separated OCR languages. Default: en")
    args = parser.parse_args()

    input_dir = Path(args.input)
    if not input_dir.is_dir():
        print(f"Error: '{input_dir}' is not a directory", file=sys.stderr)
        sys.exit(1)

    try:
        import easyocr
    except ImportError:
        print("EasyOCR is not installed. Install it with:\n    pip install easyocr",
              file=sys.stderr)
        sys.exit(1)

    images = find_images(input_dir, args.recursive, extensions=["jpg", "jpeg"])
    if not images:
        print(f"No JPG files found in '{input_dir}'"
              f"{' (searched recursively)' if args.recursive else ''}.", file=sys.stderr)
        sys.exit(1)

    print(f"Found {len(images)} photo(s).")
    print("Loading OCR model (first run downloads weights, roughly 100MB, then caches locally)...")
    reader = easyocr.Reader(args.languages.split(","), gpu=args.gpu)

    rows = []
    for i, img_path in enumerate(images, 1):
        print(f"[{i}/{len(images)}] {img_path.name}", end="\r" if sys.stdout.isatty() else "\n")
        try:
            results = reader.readtext(str(img_path))
        except Exception as e:
            rows.append([img_path.name, "error", f"OCR failed: {e}"])
            continue

        bibs = []
        for (_bbox, text, conf) in results:
            if conf < args.min_conf:
                continue
            bib = extract_bib(text, args.min_digits, args.max_digits)
            if bib:
                bibs.append(bib)

        # de-duplicate, preserve order
        seen = set()
        unique_bibs = []
        for b in bibs:
            if b not in seen:
                seen.add(b)
                unique_bibs.append(b)

        if unique_bibs:
            rows.append([img_path.name, ";".join(unique_bibs), ""])
        else:
            rows.append([img_path.name, "none", "no numeric text met the confidence/length threshold"])

    print()  # newline after progress display

    output_path = Path(args.output)
    if output_path.suffix.lower() == ".xlsx":
        write_xlsx(rows, output_path)
    else:
        write_csv(rows, output_path)

    detected = sum(1 for r in rows if r[1] not in ("none", "error"))
    print(f"Done. {detected}/{len(rows)} photos had a plausible bib number. Wrote results to {output_path}")


if __name__ == "__main__":
    main()
